"""
Excel output generation with professional formatting.
Creates a multi-sheet workbook with P&L, metrics, note breakup, validation,
and header validation for company verification.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# -------------------------------------------------------------------
# Shared styles
# -------------------------------------------------------------------

HEADER_FONT = Font(name='Arial', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='2F5496')
SECTION_FONT = Font(name='Arial', bold=True, size=10)
SECTION_FILL = PatternFill('solid', fgColor='D6E4F0')
NORMAL_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(name='Arial', size=10, bold=True)
BLUE_FONT = Font(name='Arial', size=10, color='0000FF')
GREEN_FONT = Font(name='Arial', size=10, bold=True, color='006100')
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
ORANGE_FONT = Font(name='Arial', size=10, bold=True, color='BF6900')
ORANGE_FILL = PatternFill('solid', fgColor='FFF2CC')
LIGHT_BORDER = Border(bottom=Side(style='thin', color='D9D9D9'))
TOTAL_BORDER = Border(top=Side(style='thin'), bottom=Side(style='double'))

SECTION_LABELS = {
    'pnl': 'Statement of Profit and Loss',
    'bs': 'Balance Sheet',
    'cf': 'Cash Flow Statement',
    'notes_start': 'Notes to Financial Statements',
}


def _write_header_row(ws, row: int, headers: list[str], col_start: int = 1):
    """Write a styled header row."""
    for c, h in enumerate(headers, col_start):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center' if c > col_start else 'left')


def _set_column_widths(ws, widths: dict):
    """Set column widths from a dict like {'A': 42, 'B': 20}."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# -------------------------------------------------------------------
# Sheet 1: P&L Extracted
# -------------------------------------------------------------------

PNL_ROWS = [
    ('INCOME', None, True, False),
    ('Revenue from operations', 'Revenue from operations', False, False),
    ('Other income', 'Other income', False, False),
    ('Total Income', 'Total income', False, True),
    ('', None, False, False),
    ('EXPENSES', None, True, False),
    ('Employee benefits expense', 'Employee benefits expense', False, False),
    ('Cost of professionals', 'Cost of professionals', False, False),
    ('Finance costs', 'Finance costs', False, False),
    ('Depreciation and amortisation', 'Depreciation and amortisation', False, False),
    ('Other expenses', 'Other expenses', False, False),
    ('Total Expenses', 'Total expenses', False, True),
    ('', None, False, False),
    ('Profit Before Tax (PBT)', 'Profit before tax', False, True),
    ('  Current tax', 'Current tax', False, False),
    ('  Deferred tax', 'Deferred tax', False, False),
    ('  Total tax expense', 'Total tax expense', False, False),
    ('Profit After Tax (PAT)', 'Profit for the year', False, True),
    ('', None, False, False),
    ('Total Comprehensive Income', 'Total comprehensive income', False, False),
    ('Basic EPS', 'Basic EPS', False, False),
    ('Diluted EPS', 'Diluted EPS', False, False),
]


def _write_pnl_sheet(wb: Workbook, pnl: dict, fy_current: str, fy_previous: str):
    ws = wb.active
    ws.title = "P&L - Extracted"
    _set_column_widths(ws, {'A': 42, 'B': 20, 'C': 20, 'D': 18})

    company = pnl.get('company', 'Unknown Company')
    currency = pnl.get('currency', 'INR Million')
    items = pnl.get('items', {})

    ws.merge_cells('A1:D1')
    ws['A1'] = f"{company} - Standalone P&L"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws['A2'] = f"Source: Annual Report | {currency}"
    ws['A2'].font = Font(name='Arial', size=9, italic=True, color='808080')

    r = 4
    _write_header_row(ws, r, ['Particulars', fy_current, fy_previous, 'YoY Change'])

    for label, key, is_sec, is_tot in PNL_ROWS:
        r += 1
        ws.cell(row=r, column=1, value=label)
        if is_sec:
            for c in range(1, 5):
                ws.cell(row=r, column=c).font = SECTION_FONT
                ws.cell(row=r, column=c).fill = SECTION_FILL
        elif key and key in items:
            cur = items[key].get('current', 0) or 0
            prev = items[key].get('previous', 0) or 0
            for c, v in [(2, cur), (3, prev)]:
                cell = ws.cell(row=r, column=c, value=v)
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
                cell.font = BOLD_FONT if is_tot else NORMAL_FONT
            if is_tot:
                ws.cell(row=r, column=1).font = BOLD_FONT
            ws.cell(row=r, column=4).value = f'=IF(C{r}=0,"-",(B{r}-C{r})/ABS(C{r}))'
            ws.cell(row=r, column=4).number_format = '0.0%'
            ws.cell(row=r, column=4).font = BLUE_FONT
            ws.cell(row=r, column=4).alignment = Alignment(horizontal='right')
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = TOTAL_BORDER if is_tot else LIGHT_BORDER

    ws.freeze_panes = 'A5'


# -------------------------------------------------------------------
# Sheet 2: Operating Metrics
# -------------------------------------------------------------------

METRIC_ROWS = [
    ('REVENUE', None, True, False, False),
    ('Revenue from Operations', 'Revenue from Operations', False, False, False),
    ('Other Income', 'Other Income', False, False, False),
    ('Total Income', 'Total Income', False, True, False),
    ('', None, False, False, False),
    ('OPERATING EXPENSES', None, True, False, False),
    ('Employee Benefits Expense', 'Employee Benefits Expense', False, False, False),
    ('Cost of Professionals', 'Cost of Professionals', False, False, False),
    ('Depreciation & Amortisation', 'Depreciation & Amortisation', False, False, False),
    ('Other Expenses', 'Other Expenses', False, False, False),
    ('Total Operating Expenses', 'Total Operating Expenses', False, True, False),
    ('', None, False, False, False),
    ('KEY PROFITABILITY', None, True, False, False),
    ('Operating Profit (EBIT)', 'Operating Profit (EBIT)', False, False, True),
    ('EBITDA', 'EBITDA', False, False, True),
    ('Finance Costs', 'Finance Costs', False, False, False),
    ('Profit Before Tax (PBT)', 'Profit Before Tax', False, True, False),
    ('Total Tax Expense', 'Total Tax Expense', False, False, False),
    ('Profit After Tax (PAT)', 'Profit After Tax', False, False, True),
    ('', None, False, False, False),
    ('MARGINS', None, True, False, False),
    ('Operating Margin (%)', 'Operating Margin (%)', False, False, False),
    ('EBITDA Margin (%)', 'EBITDA Margin (%)', False, False, False),
    ('PBT Margin (%)', 'PBT Margin (%)', False, False, False),
    ('PAT Margin (%)', 'PAT Margin (%)', False, False, False),
]


def _write_metrics_sheet(wb: Workbook, metrics: dict, company: str, currency: str,
                         fy_current: str, fy_previous: str):
    ws = wb.create_sheet("Operating Metrics")
    _set_column_widths(ws, {'A': 42, 'B': 20, 'C': 20, 'D': 18})

    ws.merge_cells('A1:D1')
    ws['A1'] = f"{company} - Operating Profit Analysis"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws['A2'] = f"Computed from Standalone P&L | {currency}"
    ws['A2'].font = Font(name='Arial', size=9, italic=True, color='808080')

    r = 4
    _write_header_row(ws, r, ['Particulars', fy_current, fy_previous, 'YoY Change'])

    for label, key, is_sec, is_tot, is_hl in METRIC_ROWS:
        r += 1
        ws.cell(row=r, column=1, value=label)
        if is_sec:
            for c in range(1, 5):
                ws.cell(row=r, column=c).font = SECTION_FONT
                ws.cell(row=r, column=c).fill = SECTION_FILL
        elif key and key in metrics.get('current', {}):
            cur_v = metrics['current'][key]
            prev_v = metrics['previous'][key]
            is_pct = '%' in key
            for c, v in [(2, cur_v), (3, prev_v)]:
                cell = ws.cell(row=r, column=c, value=v)
                cell.number_format = '0.00"%"' if is_pct else '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if is_hl:
                for c in range(1, 5):
                    ws.cell(row=r, column=c).font = GREEN_FONT
                    ws.cell(row=r, column=c).fill = GREEN_FILL
            elif is_tot:
                for c in [1, 2, 3]:
                    ws.cell(row=r, column=c).font = BOLD_FONT
            if is_pct:
                ws.cell(row=r, column=4).value = f'=B{r}-C{r}'
                ws.cell(row=r, column=4).number_format = '0.00" bps"'
            else:
                ws.cell(row=r, column=4).value = f'=IF(C{r}=0,"-",(B{r}-C{r})/ABS(C{r}))'
                ws.cell(row=r, column=4).number_format = '0.0%'
            ws.cell(row=r, column=4).font = BLUE_FONT
            ws.cell(row=r, column=4).alignment = Alignment(horizontal='right')
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = TOTAL_BORDER if (is_tot or is_hl) else LIGHT_BORDER

    ws.freeze_panes = 'A5'


# -------------------------------------------------------------------
# Sheet 3: Other Expenses Breakup
# -------------------------------------------------------------------

def _write_note_sheet(wb: Workbook, pnl: dict, note_items: list, note_total: dict | None,
                      note_num: str | None, fy_current: str, fy_previous: str):
    ws = wb.create_sheet("Other Expenses Breakup")
    _set_column_widths(ws, {'A': 50, 'B': 20, 'C': 20, 'D': 18, 'E': 16})

    company = pnl.get('company', 'Unknown Company')
    currency = pnl.get('currency', 'INR Million')
    items = pnl.get('items', {})

    ws.merge_cells('A1:E1')
    ws['A1'] = f"{company} - Other Expenses Breakup"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws['A2'] = f"Note {note_num or '?'} to Standalone Financial Statements | {currency}"
    ws['A2'].font = Font(name='Arial', size=9, italic=True, color='808080')

    r = 4
    _write_header_row(ws, r, ['Expense Head', fy_current, fy_previous, 'YoY Change', '% of Revenue'])

    revenue_cy = items.get('Revenue from operations', {}).get('current', 1) or 1
    pnl_total = items.get('Other expenses', {}).get('current', 0) or 0

    if note_items:
        for ni in note_items:
            r += 1
            label = ni.get('label', '')
            cur = ni.get('current', 0) or 0
            prev = ni.get('previous', 0) or 0

            is_total = (abs(cur - pnl_total) < 1) if cur and pnl_total else False
            is_sub = ' - ' in label

            ws.cell(row=r, column=1, value=f"  {label}" if is_sub else label)

            cell_b = ws.cell(row=r, column=2, value=cur)
            cell_c = ws.cell(row=r, column=3, value=prev)
            cell_b.number_format = '#,##0.00'
            cell_c.number_format = '#,##0.00'
            cell_b.alignment = Alignment(horizontal='right')
            cell_c.alignment = Alignment(horizontal='right')

            ws.cell(row=r, column=4).value = f'=IF(C{r}=0,"-",(B{r}-C{r})/ABS(C{r}))'
            ws.cell(row=r, column=4).number_format = '0.0%'
            ws.cell(row=r, column=4).font = BLUE_FONT
            ws.cell(row=r, column=4).alignment = Alignment(horizontal='right')

            ws.cell(row=r, column=5).value = cur / revenue_cy if revenue_cy else 0
            ws.cell(row=r, column=5).number_format = '0.00%'
            ws.cell(row=r, column=5).alignment = Alignment(horizontal='right')

            if is_total:
                for c in range(1, 6):
                    ws.cell(row=r, column=c).font = BOLD_FONT
                    ws.cell(row=r, column=c).border = TOTAL_BORDER
            elif is_sub:
                ws.cell(row=r, column=1).font = Font(name='Arial', size=9, italic=True, color='555555')
                cell_b.font = Font(name='Arial', size=9, color='555555')
                cell_c.font = Font(name='Arial', size=9, color='555555')
            else:
                cell_b.font = NORMAL_FONT
                cell_c.font = NORMAL_FONT

            for c in range(1, 6):
                if not is_total:
                    ws.cell(row=r, column=c).border = LIGHT_BORDER

        # Top 3 expense heads
        non_total = [ni for ni in note_items
                     if abs((ni.get('current', 0) or 0) - pnl_total) > 1]
        sorted_items = sorted(non_total, key=lambda x: abs(x.get('current', 0) or 0), reverse=True)
        top3 = sorted_items[:3]

        if top3:
            r += 2
            ws.cell(row=r, column=1, value="TOP 3 EXPENSE HEADS (by CY amount)")
            ws.cell(row=r, column=1).font = SECTION_FONT
            ws.cell(row=r, column=1).fill = ORANGE_FILL
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = ORANGE_FILL

            for t in top3:
                r += 1
                cur_val = t.get('current', 0) or 0
                ws.cell(row=r, column=1, value=f"  > {t['label']}")
                ws.cell(row=r, column=1).font = ORANGE_FONT
                ws.cell(row=r, column=2, value=cur_val)
                ws.cell(row=r, column=2).number_format = '#,##0.00'
                ws.cell(row=r, column=2).font = ORANGE_FONT
                pct = (cur_val / revenue_cy * 100) if revenue_cy else 0
                ws.cell(row=r, column=5, value=f"{pct:.2f}% of revenue")
                ws.cell(row=r, column=5).font = Font(name='Arial', size=9, italic=True, color='BF6900')

    ws.freeze_panes = 'A5'


# -------------------------------------------------------------------
# Sheet 4: Validation
# -------------------------------------------------------------------

def _write_validation_sheet(wb: Workbook, pnl: dict, metrics: dict,
                            note_total: dict | None, note_num: str | None,
                            page_headers: dict, pages: dict):
    ws = wb.create_sheet("Validation")
    _set_column_widths(ws, {'A': 55, 'B': 18, 'C': 18, 'D': 12})

    ws['A1'] = "Data Validation & Cross-Checks"
    ws['A1'].font = Font(name='Arial', bold=True, size=13, color='2F5496')

    # --- Section 1: Header Validation (Company Verification) ---
    r = 3
    ws.cell(row=r, column=1, value="STANDALONE PAGE HEADERS (verify correct company)")
    for c in range(1, 5):
        ws.cell(row=r, column=c).font = SECTION_FONT
        ws.cell(row=r, column=c).fill = SECTION_FILL

    r += 1
    _write_header_row(ws, r, ['Section', 'PDF Page', 'Header Text', ''])

    for section_key in ['pnl', 'bs', 'cf', 'notes_start']:
        page_num = pages.get(section_key)
        if page_num is None:
            continue
        r += 1
        section_name = SECTION_LABELS.get(section_key, section_key)
        header_text = page_headers.get(section_key, 'N/A')

        ws.cell(row=r, column=1, value=section_name)
        ws.cell(row=r, column=1).font = BOLD_FONT
        ws.cell(row=r, column=2, value=f"Page {page_num + 1}")
        ws.cell(row=r, column=2).font = NORMAL_FONT
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')

        # Write header text across columns C-D merged
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.cell(row=r, column=3, value=header_text)
        ws.cell(row=r, column=3).font = Font(name='Arial', size=9, color='333333')
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical='top')

        for c in range(1, 5):
            ws.cell(row=r, column=c).border = LIGHT_BORDER

    # --- Section 2: Numeric Cross-Checks ---
    r += 2
    ws.cell(row=r, column=1, value="NUMERIC CROSS-CHECKS")
    for c in range(1, 5):
        ws.cell(row=r, column=c).font = SECTION_FONT
        ws.cell(row=r, column=c).fill = SECTION_FILL

    r += 1
    _write_header_row(ws, r, ['Check', 'Computed', 'Reported', 'Status'])

    items = pnl.get('items', {})
    note_total_val = (note_total.get('current', 0) or 0) if note_total else 0
    pnl_other_exp = items.get('Other expenses', {}).get('current', 0) or 0

    checks = [
        ('Total Income - Total Expenses = PBT',
         (items.get('Total income', {}).get('current', 0) or 0) -
         (items.get('Total expenses', {}).get('current', 0) or 0),
         items.get('Profit before tax', {}).get('current', 0) or 0),
        ('PBT - Tax = PAT',
         (items.get('Profit before tax', {}).get('current', 0) or 0) -
         (items.get('Total tax expense', {}).get('current', 0) or 0),
         items.get('Profit for the year', {}).get('current', 0) or 0),
        ('OpProfit = Revenue - (Emp + CoP + Dep + OtherExp)',
         metrics['current']['Operating Profit (EBIT)'],
         metrics['current']['Revenue from Operations'] - metrics['current']['Total Operating Expenses']),
        ('EBITDA = OpProfit + Depreciation',
         metrics['current']['EBITDA'],
         metrics['current']['Operating Profit (EBIT)'] + metrics['current']['Depreciation & Amortisation']),
        (f'Note {note_num or "?"} Total = P&L Other Expenses (CY)',
         note_total_val,
         pnl_other_exp),
    ]

    for name, comp, rep in checks:
        r += 1
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=round(comp, 2)).number_format = '#,##0.00'
        ws.cell(row=r, column=3, value=round(rep, 2)).number_format = '#,##0.00'
        ok = abs(comp - rep) < 1
        ws.cell(row=r, column=4, value='PASS' if ok else 'FAIL')
        ws.cell(row=r, column=4).font = Font(
            name='Arial', bold=True,
            color='006100' if ok else 'FF0000'
        )


# -------------------------------------------------------------------
# Public API
# -------------------------------------------------------------------

def create_excel(data: dict, output_path: str) -> str:
    """
    Create a formatted Excel workbook from extracted financial data.

    Args:
        data: Dict with keys: pnl, note_items, note_total, note_number,
              fy_current, fy_previous, company, currency, pages, page_headers
        output_path: Path to save the Excel file

    Returns:
        The output path.
    """
    from app.extractor import compute_metrics

    wb = Workbook()

    pnl = data.get('pnl', {})
    company = data.get('company', pnl.get('company', 'Unknown Company'))
    currency = data.get('currency', pnl.get('currency', 'INR Million'))
    fy_current = data.get('fy_current', 'Current Year')
    fy_previous = data.get('fy_previous', 'Previous Year')
    note_items = data.get('note_items', [])
    note_total = data.get('note_total')
    note_num = data.get('note_number')
    page_headers = data.get('page_headers', {})
    pages = data.get('pages', {})

    # Ensure pnl has company/currency
    pnl['company'] = company
    pnl['currency'] = currency

    # Compute metrics
    metrics = compute_metrics(pnl)

    # Write all sheets
    _write_pnl_sheet(wb, pnl, fy_current, fy_previous)
    _write_metrics_sheet(wb, metrics, company, currency, fy_current, fy_previous)
    _write_note_sheet(wb, pnl, note_items, note_total, note_num, fy_current, fy_previous)
    _write_validation_sheet(wb, pnl, metrics, note_total, note_num, page_headers, pages)

    wb.save(output_path)
    return output_path
