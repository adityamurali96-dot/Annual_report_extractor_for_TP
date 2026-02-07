import fitz
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# UTILITY FUNCTIONS
# ============================================================

def parse_number(s):
    s = s.strip().replace(',', '').replace(' ', '')
    if s in ['-', '']:
        return 0.0
    neg = s.startswith('(') and s.endswith(')')
    if neg:
        s = s[1:-1]
    try:
        val = float(s)
        return -val if neg else val
    except:
        return None

def is_note_ref(s):
    """Note refs: integers (24, 27) or sub-notes (26.1, 4.7) - NOT values like 74.45"""
    s = s.strip()
    if re.match(r'^\d{1,2}$', s):
        return True
    if re.match(r'^\d{1,2}\.\d$', s):
        return True
    return False

def is_value_line(s):
    s = s.strip()
    if not s or s == '-':
        return True if s == '-' else False
    test = s
    if test.startswith('(') and test.endswith(')'):
        test = test[1:-1]
    test = test.replace(',', '').strip()
    try:
        float(test)
        return True
    except:
        return False


# ============================================================
# STAGE 1: TOC & Page Targeting
# ============================================================

def find_standalone_pages(pdf_path):
    doc = fitz.open(pdf_path)
    pages = {}
    for i in range(doc.page_count):
        text = doc[i].get_text()
        if 'Statement of Profit and Loss' in text and 'Standalone' in text:
            pages['pnl'] = i
        if 'Balance Sheet as at' in text and 'Standalone' in text:
            pages['bs'] = i
        if 'Cash Flow Statement' in text and 'Standalone' in text:
            pages['cf'] = i
    total = doc.page_count
    doc.close()
    return pages, total


# ============================================================
# STAGE 2: P&L Extraction
# ============================================================

def extract_pnl(pdf_path, page_idx):
    doc = fitz.open(pdf_path)
    text = doc[page_idx].get_text()
    doc.close()
    lines = [l.strip() for l in text.split('\n')]

    pnl_end = len(lines)
    for i, l in enumerate(lines):
        if l == 'ASSETS':
            pnl_end = i
            break
    pnl_lines = lines[:pnl_end]

    targets = [
        ('Revenue from operations', ['Revenue from operations']),
        ('Other income', ['Other income']),
        ('Total income', ['Total income']),
        ('Employee benefits expense', ['Employee benefits expense']),
        ('Cost of professionals', ['Cost of professionals']),
        ('Finance costs', ['Finance costs']),
        ('Depreciation and amortisation', ['Depreciation and amortisation']),
        ('Other expenses', ['Other expenses']),
        ('Total expenses', ['Total expenses']),
        ('Profit before tax', ['Profit before exceptional', 'Profit before tax']),
        ('Current tax', ['Current tax']),
        ('Deferred tax', ['Deferred tax']),
        ('Total tax expense', ['Total tax expense']),
        ('Profit for the year', ['Profit for the year']),
        ('Total comprehensive income', ['Total comprehensive income']),
        ('Basic EPS', ['Basic (In']),
        ('Diluted EPS', ['Diluted (In']),
    ]

    extracted = {}
    note_refs = {}  # Track note references for each item

    for item_name, patterns in targets:
        for i, line in enumerate(pnl_lines):
            if not any(p.lower() in line.lower() for p in patterns):
                continue
            vals = []
            note_ref = None
            for j in range(i + 1, min(i + 8, len(pnl_lines))):
                candidate = pnl_lines[j]
                if is_note_ref(candidate) and note_ref is None:
                    note_ref = candidate  # capture the note reference
                    continue
                if is_value_line(candidate):
                    vals.append(parse_number(candidate))
                    if len(vals) == 2:
                        break
                elif vals:
                    break

            if len(vals) >= 2:
                extracted[item_name] = {'current': vals[0], 'previous': vals[1]}
            elif len(vals) == 1:
                extracted[item_name] = {'current': vals[0], 'previous': 0.0}
            if note_ref:
                note_refs[item_name] = note_ref
            break

    # Detect company name from first few lines
    company = 'Unknown Company'
    for l in lines[:5]:
        if 'Limited' in l or 'Ltd' in l:
            company = l.split('—')[0].split('–')[0].strip()
            break

    return {
        'company': company,
        'currency': 'INR Million',
        'items': extracted,
        'note_refs': note_refs
    }


# ============================================================
# STAGE 2B: Generic Note Finder & Extractor
# ============================================================

def find_note_page(pdf_path, note_number, search_start_page, search_keyword="Other expenses"):
    """
    Dynamically find the PDF page containing a specific note number.
    Searches from search_start_page onwards for a heading like "27.  Other expenses"
    """
    doc = fitz.open(pdf_path)
    note_pattern = re.compile(rf'{note_number}\.\s', re.IGNORECASE)

    found_page = None
    found_line = None

    for i in range(search_start_page, doc.page_count):
        text = doc[i].get_text()
        lines = [l.strip() for l in text.split('\n')]
        for j, l in enumerate(lines):
            if note_pattern.match(l) and search_keyword.lower() in l.lower():
                found_page = i
                found_line = j
                break
        if found_page is not None:
            break

    doc.close()
    return found_page, found_line


def extract_note_breakup(pdf_path, page_idx, start_line, note_number):
    """
    Extract line items from a note breakup table.
    Reads label → CY → PY pattern until it hits the total or next note heading.
    Handles sub-items (lines starting with "- ") by prefixing parent category.
    """
    doc = fitz.open(pdf_path)
    text = doc[page_idx].get_text()
    doc.close()
    lines = [l.strip() for l in text.split('\n')]

    # Skip the note heading and column headers (For the year ended, March 31, In ₹ Million)
    data_start = start_line + 1
    while data_start < len(lines):
        l = lines[data_start].lower()
        if 'for the year' in l or 'march' in l or 'in ₹' in l or l == '':
            data_start += 1
        else:
            break

    # Parse line items
    note_items = []
    current_label = None
    parent_label = None
    i = data_start
    pending_values = []  # Track consecutive values that might be the total

    while i < len(lines):
        line = lines[i]

        # Stop conditions: next note heading or footnote
        if line and line[0].isdigit() and '.' in line[:4] and any(c.isalpha() for c in line[5:]):
            match = re.match(r'(\d+)\.', line)
            if match and match.group(1) != str(note_number):
                break

        # Stop at footnote markers (but not "- " sub-items)
        if line.startswith('*') and len(line) > 5 and any(c.isalpha() for c in line):
            break

        if is_value_line(line) or line == '-':
            val = parse_number(line) if line != '-' else 0.0
            if current_label is not None:
                existing = next((x for x in note_items if x['label'] == current_label), None)
                if existing and existing.get('previous') is None:
                    existing['previous'] = val
                    current_label = None  # Both CY/PY filled, reset
                elif existing is None:
                    note_items.append({'label': current_label, 'current': val, 'previous': None})
            else:
                # Value line with no current label - could be the TOTAL row
                pending_values.append(val)
        else:
            if line:
                pending_values = []  # Reset if we hit a label
                if line.startswith('- '):
                    current_label = f"{parent_label} - {line[2:]}" if parent_label else line[2:]
                else:
                    current_label = line
                    parent_label = line

        i += 1

    # Clean up
    result = []
    for item in note_items:
        if item.get('previous') is None:
            item['previous'] = 0.0
        result.append(item)

    # The pending_values at the end (2 consecutive values with no label) are the total
    total_item = None
    if len(pending_values) >= 2:
        total_item = {'label': 'Total Other Expenses', 'current': pending_values[0], 'previous': pending_values[1]}
    elif result:
        # Fallback: last item might be the total
        total_item = result[-1]

    return result, total_item


# ============================================================
# STAGE 3: Compute Operating Profit & Metrics
# ============================================================

def compute_metrics(pnl):
    items = pnl['items']
    metrics = {}
    for period in ['current', 'previous']:
        rev = items.get('Revenue from operations', {}).get(period, 0)
        oi = items.get('Other income', {}).get(period, 0)
        emp = items.get('Employee benefits expense', {}).get(period, 0)
        cop = items.get('Cost of professionals', {}).get(period, 0)
        dep = items.get('Depreciation and amortisation', {}).get(period, 0)
        oe = items.get('Other expenses', {}).get(period, 0)
        fc = items.get('Finance costs', {}).get(period, 0)
        tax = items.get('Total tax expense', {}).get(period, 0)
        pat = items.get('Profit for the year', {}).get(period, 0)
        pbt = items.get('Profit before tax', {}).get(period, 0)

        opex = emp + cop + dep + oe
        op_profit = rev - opex
        ebitda = op_profit + dep

        metrics[period] = {
            'Revenue from Operations': rev,
            'Other Income': oi,
            'Total Income': rev + oi,
            'Employee Benefits Expense': emp,
            'Cost of Professionals': cop,
            'Depreciation & Amortisation': dep,
            'Other Expenses': oe,
            'Total Operating Expenses': opex,
            'Operating Profit (EBIT)': op_profit,
            'Finance Costs': fc,
            'Profit Before Tax': pbt,
            'Total Tax Expense': tax,
            'Profit After Tax': pat,
            'EBITDA': ebitda,
            'Operating Margin (%)': (op_profit / rev * 100) if rev else 0,
            'EBITDA Margin (%)': (ebitda / rev * 100) if rev else 0,
            'PBT Margin (%)': (pbt / rev * 100) if rev else 0,
            'PAT Margin (%)': (pat / rev * 100) if rev else 0,
        }
    return metrics


# ============================================================
# STAGE 4: Excel Output
# ============================================================

def create_excel(metrics, pnl, note_items, note_total, output_path):
    wb = Workbook()

    # Styles
    hf = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='2F5496')
    sf = Font(name='Arial', bold=True, size=10)
    sfill = PatternFill('solid', fgColor='D6E4F0')
    nf = Font(name='Arial', size=10)
    bf = Font(name='Arial', size=10, bold=True)
    blf = Font(name='Arial', size=10, color='0000FF')
    gf = Font(name='Arial', size=10, bold=True, color='006100')
    gfill = PatternFill('solid', fgColor='C6EFCE')
    orange_font = Font(name='Arial', size=10, bold=True, color='BF6900')
    orange_fill = PatternFill('solid', fgColor='FFF2CC')
    lb = Border(bottom=Side(style='thin', color='D9D9D9'))
    tb = Border(top=Side(style='thin'), bottom=Side(style='double'))

    items = pnl['items']
    company = pnl['company']

    # ---- Sheet 1: Extracted P&L ----
    ws = wb.active
    ws.title = "P&L - Extracted"
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18

    ws.merge_cells('A1:D1')
    ws['A1'] = f"{company} — Standalone P&L"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws['A2'] = f"Source: Annual Report | {pnl['currency']}"
    ws['A2'].font = Font(name='Arial', size=9, italic=True, color='808080')

    r = 4
    for c, h in enumerate(['Particulars', 'FY 2024-25', 'FY 2023-24', 'YoY Change'], 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = hf; cell.fill = hfill
        cell.alignment = Alignment(horizontal='center' if c > 1 else 'left')

    rows_def = [
        ('INCOME', None, True, False),
        ('Revenue from operations', 'Revenue from operations', False, False),
        ('Other income', 'Other income', False, False),
        ('Total Income', 'Total income', False, True),
        ('', None, False, False),
        ('EXPENSES', None, True, False),
        ('Employee benefits expense', 'Employee benefits expense', False, False),
        ('Cost of professionals', 'Cost of professionals', False, False),
        ('Finance costs', 'Finance costs', False, False),
        ('Depreciation and amortisation', 'Depreciation and amortisation', False, False),
        ('Other expenses', 'Other expenses', False, False),
        ('Total Expenses', 'Total expenses', False, True),
        ('', None, False, False),
        ('Profit Before Tax (PBT)', 'Profit before tax', False, True),
        ('  Current tax', 'Current tax', False, False),
        ('  Deferred tax', 'Deferred tax', False, False),
        ('  Total tax expense', 'Total tax expense', False, False),
        ('Profit After Tax (PAT)', 'Profit for the year', False, True),
        ('', None, False, False),
        ('Basic EPS (₹)', 'Basic EPS', False, False),
        ('Diluted EPS (₹)', 'Diluted EPS', False, False),
    ]

    r = 4
    for label, key, is_sec, is_tot in rows_def:
        r += 1
        ws.cell(row=r, column=1, value=label)
        if is_sec:
            for c in range(1, 5):
                ws.cell(row=r, column=c).font = sf
                ws.cell(row=r, column=c).fill = sfill
        elif key and key in items:
            cur, prev = items[key]['current'], items[key]['previous']
            for c, v in [(2, cur), (3, prev)]:
                cell = ws.cell(row=r, column=c, value=v)
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
                cell.font = bf if is_tot else nf
            if is_tot:
                ws.cell(row=r, column=1).font = bf
            ws.cell(row=r, column=4).value = f'=IF(C{r}=0,"-",(B{r}-C{r})/ABS(C{r}))'
            ws.cell(row=r, column=4).number_format = '0.0%'
            ws.cell(row=r, column=4).font = blf
            ws.cell(row=r, column=4).alignment = Alignment(horizontal='right')
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = tb if is_tot else lb

    # ---- Sheet 2: Operating Metrics ----
    ws2 = wb.create_sheet("Operating Metrics")
    ws2.column_dimensions['A'].width = 42
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 18

    ws2.merge_cells('A1:D1')
    ws2['A1'] = f"{company} — Operating Profit Analysis"
    ws2['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws2['A2'] = f"Computed from Standalone P&L | {pnl['currency']}"
    ws2['A2'].font = Font(name='Arial', size=9, italic=True, color='808080')

    r = 4
    for c, h in enumerate(['Particulars', 'FY 2024-25', 'FY 2023-24', 'YoY Change'], 1):
        cell = ws2.cell(row=r, column=c, value=h)
        cell.font = hf; cell.fill = hfill
        cell.alignment = Alignment(horizontal='center' if c > 1 else 'left')

    m_rows = [
        ('REVENUE', None, True, False, False),
        ('Revenue from Operations', 'Revenue from Operations', False, False, False),
        ('Other Income', 'Other Income', False, False, False),
        ('Total Income', 'Total Income', False, True, False),
        ('', None, False, False, False),
        ('OPERATING EXPENSES', None, True, False, False),
        ('Employee Benefits Expense', 'Employee Benefits Expense', False, False, False),
        ('Cost of Professionals', 'Cost of Professionals', False, False, False),
        ('Depreciation & Amortisation', 'Depreciation & Amortisation', False, False, False),
        ('Other Expenses', 'Other Expenses', False, False, False),
        ('Total Operating Expenses', 'Total Operating Expenses', False, True, False),
        ('', None, False, False, False),
        ('KEY PROFITABILITY', None, True, False, False),
        ('Operating Profit (EBIT)', 'Operating Profit (EBIT)', False, False, True),
        ('EBITDA', 'EBITDA', False, False, True),
        ('Finance Costs', 'Finance Costs', False, False, False),
        ('Profit Before Tax (PBT)', 'Profit Before Tax', False, True, False),
        ('Total Tax Expense', 'Total Tax Expense', False, False, False),
        ('Profit After Tax (PAT)', 'Profit After Tax', False, False, True),
        ('', None, False, False, False),
        ('MARGINS', None, True, False, False),
        ('Operating Margin (%)', 'Operating Margin (%)', False, False, False),
        ('EBITDA Margin (%)', 'EBITDA Margin (%)', False, False, False),
        ('PBT Margin (%)', 'PBT Margin (%)', False, False, False),
        ('PAT Margin (%)', 'PAT Margin (%)', False, False, False),
    ]

    r = 4
    for label, key, is_sec, is_tot, is_hl in m_rows:
        r += 1
        ws2.cell(row=r, column=1, value=label)
        if is_sec:
            for c in range(1, 5):
                ws2.cell(row=r, column=c).font = sf
                ws2.cell(row=r, column=c).fill = sfill
        elif key and key in metrics.get('current', {}):
            cur_v = metrics['current'][key]
            prev_v = metrics['previous'][key]
            is_pct = '%' in key
            for c, v in [(2, cur_v), (3, prev_v)]:
                cell = ws2.cell(row=r, column=c, value=v)
                cell.number_format = '0.00"%"' if is_pct else '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if is_hl:
                for c in range(1, 5):
                    ws2.cell(row=r, column=c).font = gf
                    ws2.cell(row=r, column=c).fill = gfill
            elif is_tot:
                for c in [1, 2, 3]:
                    ws2.cell(row=r, column=c).font = bf
            if is_pct:
                ws2.cell(row=r, column=4).value = f'=B{r}-C{r}'
                ws2.cell(row=r, column=4).number_format = '0.00" bps"'
            else:
                ws2.cell(row=r, column=4).value = f'=IF(C{r}=0,"-",(B{r}-C{r})/ABS(C{r}))'
                ws2.cell(row=r, column=4).number_format = '0.0%'
            ws2.cell(row=r, column=4).font = blf
            ws2.cell(row=r, column=4).alignment = Alignment(horizontal='right')
        for c in range(1, 5):
            ws2.cell(row=r, column=c).border = tb if (is_tot or is_hl) else lb

    # ---- Sheet 3: Other Expenses Breakup ----
    ws3 = wb.create_sheet("Other Expenses Breakup")
    ws3.column_dimensions['A'].width = 50
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 20
    ws3.column_dimensions['D'].width = 18
    ws3.column_dimensions['E'].width = 16

    ws3.merge_cells('A1:E1')
    ws3['A1'] = f"{company} — Other Expenses Breakup"
    ws3['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')

    note_num = pnl.get('note_refs', {}).get('Other expenses', '?')
    ws3['A2'] = f"Note {note_num} to Standalone Financial Statements | {pnl['currency']}"
    ws3['A2'].font = Font(name='Arial', size=9, italic=True, color='808080')

    r = 4
    for c, h in enumerate(['Expense Head', 'FY 2024-25', 'FY 2023-24', 'YoY Change', '% of Revenue'], 1):
        cell = ws3.cell(row=r, column=c, value=h)
        cell.font = hf; cell.fill = hfill
        cell.alignment = Alignment(horizontal='center' if c > 1 else 'left')

    # Write note items
    revenue_cy = items.get('Revenue from operations', {}).get('current', 1)
    total_row = None

    if note_items:
        for ni in note_items:
            r += 1
            label = ni['label']
            cur = ni['current']
            prev = ni['previous']

            # Check if this is the total line (last item matching P&L total)
            pnl_total = items.get('Other expenses', {}).get('current', 0)
            is_total = (abs(cur - pnl_total) < 1) if cur else False

            # Sub-items get indented
            is_sub = ' - ' in label
            display_label = f"  {label}" if is_sub else label

            ws3.cell(row=r, column=1, value=display_label)

            cell_b = ws3.cell(row=r, column=2, value=cur)
            cell_c = ws3.cell(row=r, column=3, value=prev)
            cell_b.number_format = '#,##0.00'
            cell_c.number_format = '#,##0.00'
            cell_b.alignment = Alignment(horizontal='right')
            cell_c.alignment = Alignment(horizontal='right')

            # YoY Change
            ws3.cell(row=r, column=4).value = f'=IF(C{r}=0,"-",(B{r}-C{r})/ABS(C{r}))'
            ws3.cell(row=r, column=4).number_format = '0.0%'
            ws3.cell(row=r, column=4).font = blf
            ws3.cell(row=r, column=4).alignment = Alignment(horizontal='right')

            # % of Revenue
            ws3.cell(row=r, column=5).value = cur / revenue_cy if revenue_cy else 0
            ws3.cell(row=r, column=5).number_format = '0.00%'
            ws3.cell(row=r, column=5).alignment = Alignment(horizontal='right')

            if is_total:
                total_row = r
                for c in range(1, 6):
                    ws3.cell(row=r, column=c).font = bf
                    ws3.cell(row=r, column=c).border = tb
                # Highlight top 3 expenses
            elif is_sub:
                ws3.cell(row=r, column=1).font = Font(name='Arial', size=9, italic=True, color='555555')
                cell_b.font = Font(name='Arial', size=9, color='555555')
                cell_c.font = Font(name='Arial', size=9, color='555555')
            else:
                cell_b.font = nf
                cell_c.font = nf

            for c in range(1, 6):
                if not is_total:
                    ws3.cell(row=r, column=c).border = lb

        # Highlight top expenses by adding a note
        if note_items:
            non_total = [ni for ni in note_items if abs(ni['current'] - items.get('Other expenses', {}).get('current', 0)) > 1]
            sorted_items = sorted(non_total, key=lambda x: abs(x['current']), reverse=True)
            top3 = sorted_items[:3] if len(sorted_items) >= 3 else sorted_items

            r += 2
            ws3.cell(row=r, column=1, value="TOP 3 EXPENSE HEADS (by CY amount)")
            ws3.cell(row=r, column=1).font = sf
            ws3.cell(row=r, column=1).fill = orange_fill
            for c in range(2, 6):
                ws3.cell(row=r, column=c).fill = orange_fill

            for t in top3:
                r += 1
                ws3.cell(row=r, column=1, value=f"  ▸ {t['label']}")
                ws3.cell(row=r, column=1).font = orange_font
                ws3.cell(row=r, column=2, value=t['current'])
                ws3.cell(row=r, column=2).number_format = '#,##0.00'
                ws3.cell(row=r, column=2).font = orange_font
                pct = (t['current'] / revenue_cy * 100) if revenue_cy else 0
                ws3.cell(row=r, column=5, value=f"{pct:.2f}% of revenue")
                ws3.cell(row=r, column=5).font = Font(name='Arial', size=9, italic=True, color='BF6900')

    ws3.freeze_panes = 'A5'

    # ---- Sheet 4: Validation ----
    ws4 = wb.create_sheet("Validation")
    ws4.column_dimensions['A'].width = 55
    ws4.column_dimensions['B'].width = 18
    ws4.column_dimensions['C'].width = 18
    ws4.column_dimensions['D'].width = 12

    ws4['A1'] = "Data Validation & Cross-Checks"
    ws4['A1'].font = Font(name='Arial', bold=True, size=13, color='2F5496')

    r = 3
    for c, h in enumerate(['Check', 'Computed', 'Reported', 'Status'], 1):
        ws4.cell(row=r, column=c, value=h).font = hf
        ws4.cell(row=r, column=c).fill = hfill

    # Note total validation
    note_total_val = note_total['current'] if note_total else 0
    pnl_other_exp = items.get('Other expenses', {}).get('current', 0)

    checks = [
        ('Total Income - Total Expenses = PBT',
         items.get('Total income', {}).get('current', 0) - items.get('Total expenses', {}).get('current', 0),
         items.get('Profit before tax', {}).get('current', 0)),
        ('PBT - Tax = PAT',
         items.get('Profit before tax', {}).get('current', 0) - items.get('Total tax expense', {}).get('current', 0),
         items.get('Profit for the year', {}).get('current', 0)),
        ('OpProfit = Revenue - (Emp + CoP + Dep + OtherExp)',
         metrics['current']['Operating Profit (EBIT)'],
         metrics['current']['Revenue from Operations'] - metrics['current']['Total Operating Expenses']),
        ('EBITDA = OpProfit + Depreciation',
         metrics['current']['EBITDA'],
         metrics['current']['Operating Profit (EBIT)'] + metrics['current']['Depreciation & Amortisation']),
        (f'Note {note_num} Total = P&L Other Expenses (CY)',
         note_total_val,
         pnl_other_exp),
    ]

    for name, comp, rep in checks:
        r += 1
        ws4.cell(row=r, column=1, value=name)
        ws4.cell(row=r, column=2, value=round(comp, 2)).number_format = '#,##0.00'
        ws4.cell(row=r, column=3, value=round(rep, 2)).number_format = '#,##0.00'
        ok = abs(comp - rep) < 1
        ws4.cell(row=r, column=4, value='✓ PASS' if ok else '✗ FAIL')
        ws4.cell(row=r, column=4).font = Font(name='Arial', bold=True, color='006100' if ok else 'FF0000')

    ws.freeze_panes = 'A5'
    ws2.freeze_panes = 'A5'
    wb.save(output_path)
    return output_path


# ============================================================
# MAIN
# ============================================================

pdf_path = "/mnt/user-data/uploads/659e0149-a77e-46a1-adab-05e676fe9996.pdf"

print("STAGE 1: Page Targeting")
pages, total = find_standalone_pages(pdf_path)
print(f"Standalone pages: {pages} (total: {total})")

print("\nSTAGE 2A: Extract P&L")
pnl = extract_pnl(pdf_path, pages['pnl'])
print(f"Company: {pnl['company']}")
print(f"Note refs: {pnl['note_refs']}")
for k, v in pnl['items'].items():
    print(f"  {k:42s} | CY: {v['current']:>14,.2f} | PY: {v['previous']:>14,.2f}")

print("\nSTAGE 2B: Find & Extract Other Expenses Note")
note_num = pnl['note_refs'].get('Other expenses', '27')
print(f"Other Expenses note ref from P&L: {note_num}")

# Search for the note in standalone notes section (after the financial statements)
note_page, note_line = find_note_page(pdf_path, note_num, pages['pnl'], "Other expenses")
print(f"Found Note {note_num} on PDF page {note_page + 1} (0-idx: {note_page}), line {note_line}")

note_items, note_total = extract_note_breakup(pdf_path, note_page, note_line, note_num)
print(f"\nExtracted {len(note_items)} line items:")
for ni in note_items:
    print(f"  {ni['label']:50s} | CY: {ni['current']:>12,.2f} | PY: {ni['previous']:>12,.2f}")
if note_total:
    print(f"\nTotal: CY {note_total['current']:,.2f} | PY {note_total['previous']:,.2f}")
    pnl_total = pnl['items'].get('Other expenses', {}).get('current', 0)
    print(f"P&L Other Expenses: {pnl_total:,.2f}")
    print(f"Match: {'✓' if abs(note_total['current'] - pnl_total) < 1 else '✗'}")

print("\nSTAGE 3: Compute Metrics")
metrics = compute_metrics(pnl)
for key in ['Revenue from Operations', 'Operating Profit (EBIT)', 'EBITDA',
            'Profit After Tax', 'Operating Margin (%)', 'EBITDA Margin (%)', 'PAT Margin (%)']:
    v = metrics['current'][key]
    print(f"  {key:42s} | {v:>10.2f}{'%' if '%' in key else ''}")

print("\nSTAGE 4: Excel Output")
out = create_excel(metrics, pnl, note_items, note_total,
                   "/home/claude/persistent_standalone_financials.xlsx")
print(f"Saved: {out}")
